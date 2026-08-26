# -*- coding: utf-8 -*-
"""
Sincroniza los links de Power BI del tablero contra la API, sin tocar los links
cargados a mano.

Como funciona
-------------
El Excel pasa a tener dos hojas:

  Tableros   Lo que lee index.html. Columna Origen: MANUAL (lo cargas vos) o API
             (lo genera este script). Las filas MANUAL no se tocan nunca.
  Catalogo   Todos los reportes que existen en las areas de trabajo de Power BI,
             con una columna Publicar. Solo los SI bajan a Tableros.

Reglas
------
* Publicar vacio  -> no se publica, pero queda listado para que lo tildes.
* Publicar NO     -> no se publica y no vuelve a proponerse.
* Nombre, Agrupacion y Descripcion son tuyos: se preservan por reportId aunque
  el reporte se renombre en Power BI.
* Series mensuales: 'COMPARATIVO JUNIO 2026' / 'Informe Hctos - Junio 26' se
  publican como 'COMPARATIVO 06-2026' / 'HECTOLITROS 06-2026', que es lo que
  index.html colapsa en un desplegable de meses. Un mes nuevo de una serie que
  ya tiene meses publicados se marca SI automaticamente.

Uso
---
  python sincronizar_powerbi.py            # actualiza "LINKS POWER BI.xlsx"
  python sincronizar_powerbi.py --dry-run  # solo informa, no escribe nada
  python sincronizar_powerbi.py --test     # trabaja sobre el _test, no toca produccion

Auth: reutiliza la sesion device-code del orquestador
(Scripts/orquestador/.pbi_token.json). Si el refresh token vencio, hay que
volver a loguear desde el orquestador (boton de Power BI).
"""
import argparse
import json
import os
import re
import sys
import time
import unicodedata
import urllib.parse as _up
import urllib.request as _ur

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

AQUI = os.path.dirname(os.path.abspath(__file__))
CONFIG = os.path.join(AQUI, "powerbi_areas.json")
EXCEL = "LINKS POWER BI.xlsx"            # el que lee index.html
EXCEL_TEST = "LINKS POWER BI _test.xlsx"  # banco de pruebas, no se publica
TOKEN_ORQUESTADOR = r"H:\Automatizacion de reportes\Scripts\orquestador\.pbi_token.json"

PBI_CLIENT_ID = "1db21c69-6970-4590-ba27-6b41cd555b4e"
PBI_TENANT_ID = "2c7a392d-4df9-434a-859c-2f6bb1519f7a"
PBI_SCOPE = ("https://analysis.windows.net/powerbi/api/.default "
             "offline_access openid profile")
PBI_AUTORIDAD = "https://login.microsoftonline.com/%s/oauth2/v2.0" % PBI_TENANT_ID
PBI_API = "https://api.powerbi.com/v1.0/myorg"
# La API de Fabric es la unica que informa en que subcarpeta vive un reporte.
FABRIC_API = "https://api.fabric.microsoft.com/v1"

# Formato exacto de los links que ya venias usando (secure embed, sin edicion).
PLANTILLA_LINK = "https://app.powerbi.com/reportEmbed?reportId=%s&autoAuth=true"


def link_de(rid, extra=""):
    """Link secure-embed del informe, con los parametros extra de la config."""
    url = PLANTILLA_LINK % rid
    extra = str(extra or "").strip().lstrip("&?")
    return url + "&" + extra if extra else url

COLS_TABLEROS = ["Area", "Agrupacion", "Nombre", "Link", "Descripcion", "Origen",
                 "Tipo",    # INFORME o APP: define el boton y el conteo
                 "Vistas"]  # vistas de los ultimos meses, para ordenar
COLS_CATALOGO = ["Publicar", "Area", "Agrupacion", "Nombre", "Descripcion",
                 "Serie", "Mes", "NombrePowerBI", "Carpeta", "AreaDeTrabajo",
                 "reportId", "Vistas"]

MESES_NOMBRE = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}


class NoLogueado(Exception):
    pass


# ---------------------------------------------------------------- utilidades

def norm(s):
    """minusculas, sin acentos, espacios colapsados."""
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().lower()


def id_de_link(link):
    m = re.search(r"reportId=([0-9a-fA-F-]{36})", str(link or ""))
    return m.group(1).lower() if m else None


# --------------------------------------------------------------------- auth

def _post_form(url, campos):
    datos = _up.urlencode(campos).encode("utf-8")
    req = _ur.Request(url, data=datos, method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")
    with _ur.urlopen(req, timeout=30) as r:
        return json.loads(r.read().decode("utf-8"))


_TOKEN = [None]


def access_token():
    """Devuelve un access_token valido, renovando y regrabando el cache del
    orquestador igual que el orquestador (para no cortarle la cadena)."""
    if _TOKEN[0]:
        return _TOKEN[0]
    if not os.path.isfile(TOKEN_ORQUESTADOR):
        raise NoLogueado("no existe %s" % TOKEN_ORQUESTADOR)
    with open(TOKEN_ORQUESTADOR, encoding="utf-8") as f:
        cache = json.load(f)
    if cache.get("access_token") and cache.get("expira_en", 0) > time.time() + 60:
        _TOKEN[0] = cache["access_token"]
        return _TOKEN[0]
    rt = cache.get("refresh_token")
    if not rt:
        raise NoLogueado("no hay refresh_token en el cache")
    try:
        tok = _post_form(PBI_AUTORIDAD + "/token", {
            "grant_type": "refresh_token",
            "client_id": PBI_CLIENT_ID,
            "refresh_token": rt,
            "scope": PBI_SCOPE,
        })
    except Exception as e:  # noqa: BLE001
        raise NoLogueado("fallo el refresh: %s" % e)
    if not tok.get("access_token"):
        raise NoLogueado("el refresh no devolvio access_token")
    cache["access_token"] = tok["access_token"]
    cache["expira_en"] = time.time() + int(tok.get("expires_in", 3600)) - 120
    if tok.get("refresh_token"):
        cache["refresh_token"] = tok["refresh_token"]
    with open(TOKEN_ORQUESTADOR, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)
    _TOKEN[0] = cache["access_token"]
    return _TOKEN[0]


def api(ruta):
    req = _ur.Request(PBI_API + ruta, method="GET")
    req.add_header("Authorization", "Bearer " + access_token())
    with _ur.urlopen(req, timeout=60) as r:
        return json.loads(r.read().decode("utf-8"))


def api_fabric(ruta):
    req = _ur.Request(FABRIC_API + ruta, method="GET")
    req.add_header("Authorization", "Bearer " + access_token())
    with _ur.urlopen(req, timeout=60) as r:
        return json.loads(r.read().decode("utf-8"))


def carpetas_por_reporte(ws_id):
    """reportId -> nombre de la subcarpeta donde vive (o '' si esta en la raiz).

    La API clasica de Power BI no dice nada de subcarpetas; hay que ir a la de
    Fabric. Sirve de respaldo para detectar un informe mensual que quedo sin el
    mes en el nombre pero esta guardado en la carpeta del mes.
    Si la API de Fabric no responde, devuelve {} y se sigue solo por nombre.
    """
    try:
        nombres = {f["id"]: (f.get("displayName") or "")
                   for f in api_fabric("/workspaces/%s/folders" % ws_id)["value"]}
        items = api_fabric("/workspaces/%s/items?type=Report" % ws_id)["value"]
    except Exception:  # noqa: BLE001
        return {}
    return {(it.get("id") or "").lower(): nombres.get(it.get("folderId"), "")
            for it in items}


def api_post(ruta, cuerpo):
    datos = json.dumps(cuerpo).encode("utf-8")
    req = _ur.Request(PBI_API + ruta, data=datos, method="POST")
    req.add_header("Authorization", "Bearer " + access_token())
    req.add_header("Content-Type", "application/json")
    with _ur.urlopen(req, timeout=120) as r:
        return json.loads(r.read().decode("utf-8"))


def _guid(v):
    """Users[UserGuid] viene como '{ABC-...}' y Views[UserGuid] como 'ABC-...'.
    Se comparan sin llaves ni mayusculas."""
    return str(v or "").strip().strip("{}").lower()


def vistas_por_reporte(ws_id, excluir_upn=()):
    """(reportId -> vistas, vistas_descartadas)

    Lee el modelo de Usage Metrics del workspace. Power BI lo crea recien cuando
    alguien abre "Metricas de uso" desde el Service en esa area de trabajo; si no
    existe devuelve ({}, 0) y el orden pasa a ser alfabetico.

    excluir_upn: cuentas cuyas vistas no cuentan (las propias, tipicamente, que
    si no inflan justo los informes que uno mas toca al desarrollarlos). Se
    compara por texto contenido en el UserPrincipalName, sin mayusculas.
    """
    try:
        ds = api("/groups/%s/datasets" % ws_id)["value"]
    except Exception:  # noqa: BLE001
        return {}, 0
    um = [d for d in ds if "usage metrics" in norm(d.get("name"))]
    if not um:
        return {}, 0
    ruta = "/groups/%s/datasets/%s/executeQueries" % (ws_id, um[0]["id"])

    def consultar(dax):
        res = api_post(ruta, {"queries": [{"query": dax}]})
        return res["results"][0]["tables"][0]["rows"]

    # que UserGuid corresponde a las cuentas a descartar
    fuera = set()
    if excluir_upn:
        try:
            for f in consultar("EVALUATE SELECTCOLUMNS(Users, "
                               "\"upn\", Users[UserPrincipalName], "
                               "\"guid\", Users[UserGuid])"):
                upn = norm(f.get("[upn]"))
                if any(x in upn for x in excluir_upn):
                    fuera.add(_guid(f.get("[guid]")))
        except Exception:  # noqa: BLE001
            pass   # sin tabla de usuarios se cuenta todo, mejor que fallar

    # se agrupa tambien por usuario para poder descartar filas
    try:
        filas = consultar("EVALUATE SUMMARIZECOLUMNS(Views[ReportGuid], "
                          "Views[UserGuid], "
                          "\"vistas\", SUM(Views[GranularViewsCount]))")
    except Exception:  # noqa: BLE001
        return {}, 0
    out, descartadas = {}, 0
    for f in filas:
        rid = str(f.get("Views[ReportGuid]") or "").lower()
        try:
            v = int(f.get("[vistas]") or 0)
        except (TypeError, ValueError):
            continue
        if _guid(f.get("Views[UserGuid]")) in fuera:
            descartadas += v
            continue
        out[rid] = out.get(rid, 0) + v
    return out, descartadas


# ------------------------------------------------------------ series mensuales

_RE_MES_TEXTO = re.compile(
    r"^(?P<base>.*?)[\s\-\u2013]*\b(?P<mes>%s)\s+(?P<anio>\d{4}|\d{2})\s*$"
    % "|".join(MESES_NOMBRE), re.IGNORECASE)
_RE_MES_NUM = re.compile(r"^(?P<base>.*?)\s+(?P<mm>\d{2})-(?P<anio>\d{4})\s*$")


def partir_periodo(texto):
    """Separa un texto en (lo que sobra, (anio, mes)).
    'COMPARATIVO JUNIO 2026' -> ('COMPARATIVO', (2026, 6))
    'Junio - 25'             -> ('',           (2025, 6))
    Si no encuentra un mes devuelve (texto, None)."""
    texto = str(texto or "")
    m = _RE_MES_TEXTO.match(texto)
    if m:
        mes = MESES_NOMBRE[norm(m.group("mes"))]
        anio = int(m.group("anio"))
    else:
        m = _RE_MES_NUM.match(texto)
        if not m:
            return texto.strip(), None
        mes = int(m.group("mm"))
        anio = int(m.group("anio"))
    if anio < 100:
        anio += 2000
    base = re.sub(r"[\s\-\u2013]+$", "", m.group("base")).strip()
    return base, (anio, mes)


def partir_serie(nombre_pbi, alias, carpeta=""):
    """('COMPARATIVO JUNIO 2026') -> ('COMPARATIVO', (2026, 6))

    Primero busca el mes en el nombre del reporte. Si el nombre no lo dice,
    lo busca en el nombre de la subcarpeta ('Agosto - 26'): ahi la serie es el
    nombre completo del reporte. Si no es mensual devuelve (None, None)."""
    base, per = partir_periodo(nombre_pbi)
    if per is None:
        # respaldo: el mes lo pone la carpeta
        _, per_carp = partir_periodo(carpeta)
        if per_carp is None:
            return None, None
        base, per = str(nombre_pbi or "").strip(), per_carp
    if not base:
        return None, None
    serie = alias.get(norm(base), base.upper())
    return serie, per


def agrupacion_de(nombre_pbi, reglas):
    """Primera regla cuyo texto aparezca en el nombre del reporte. '' si ninguna."""
    n = norm(nombre_pbi)
    for aguja, agrupacion in reglas:
        if aguja and aguja in n:
            return agrupacion
    return ""


def nombre_publicado(serie, per):
    return "%s %02d-%04d" % (serie, per[1], per[0])


# ----------------------------------------------------------------- excel: leer

def leer_hoja(path, titulos_posibles, cols):
    """Devuelve lista de dicts de la primera hoja cuyo titulo matchee.
    Si no existe la hoja, lista vacia."""
    if not os.path.isfile(path):
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    hoja = None
    for t in titulos_posibles:
        for ws in wb.worksheets:
            if norm(ws.title) == norm(t):
                hoja = ws
                break
        if hoja:
            break
    if hoja is None:
        return []
    filas = list(hoja.iter_rows(values_only=True))
    if not filas:
        return []
    cab = [norm(c) for c in filas[0]]
    idx = {}
    for c in cols:
        n = norm(c)
        idx[c] = cab.index(n) if n in cab else None
    out = []
    for f in filas[1:]:
        d = {}
        for c in cols:
            i = idx[c]
            v = f[i] if (i is not None and i < len(f)) else None
            d[c] = "" if v is None else str(v).strip()
        if any(d.values()):
            out.append(d)
    return out


def leer_tableros_legacy(path):
    """La hoja principal del Excel (con o sin columna Origen)."""
    return leer_hoja(path, ["Tableros", "Hoja1", "Sheet1"], COLS_TABLEROS)


# ---------------------------------------------------------------- excel: escribir

FILL_CAB = PatternFill("solid", fgColor="1B4FE5")
FILL_SI = PatternFill("solid", fgColor="E7F4EC")
FILL_NUEVO = PatternFill("solid", fgColor="FFF4CE")

ANCHOS_TABLEROS = {"Link": 62, "Nombre": 38, "Descripcion": 45, "Area": 20,
                   "Agrupacion": 24, "Origen": 9, "Tipo": 9, "Vistas": 8}
ANCHOS_CATALOGO = {"Publicar": 10, "Area": 14, "Agrupacion": 24, "Nombre": 36,
                   "Descripcion": 40, "Serie": 26, "Mes": 10,
                   "NombrePowerBI": 40, "Carpeta": 16, "AreaDeTrabajo": 26,
                   "reportId": 38, "Vistas": 8}


def _volcar(ws, cols, filas, anchos, validar_publicar=False):
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cel = ws.cell(1, c)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = FILL_CAB
        cel.alignment = Alignment(vertical="center")
    for f in filas:
        ws.append([f.get(c, "") for c in cols])
    for c, nombre in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c)].width = anchos.get(nombre, 18)
    ws.freeze_panes = "A2"
    ultima = max(2, len(filas) + 1)
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), ultima)
    if validar_publicar:
        dv = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
        dv.promptTitle = "Publicar"
        dv.prompt = "SI publica en el tablero. NO lo descarta. Vacio = no publica."
        ws.add_data_validation(dv)
        dv.add("A2:A%d" % ultima)


def escribir(path, tableros, catalogo, nuevos_ids):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tableros"
    _volcar(ws, COLS_TABLEROS, tableros, ANCHOS_TABLEROS)
    wc = wb.create_sheet("Catalogo")
    _volcar(wc, COLS_CATALOGO, catalogo, ANCHOS_CATALOGO, validar_publicar=True)
    # resaltado: verde lo publicado, amarillo lo recien descubierto
    for i, f in enumerate(catalogo, start=2):
        if norm(f["Publicar"]) == "si":
            wc.cell(i, 1).fill = FILL_SI
        elif f["reportId"] in nuevos_ids:
            for c in range(1, len(COLS_CATALOGO) + 1):
                wc.cell(i, c).fill = FILL_NUEVO
    wb.save(path)


# ------------------------------------------------------------------ principal

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--entrada", default=EXCEL,
                    help="Excel a leer.")
    ap.add_argument("--salida", default=EXCEL,
                    help="Excel a escribir.")
    ap.add_argument("--test", action="store_true",
                    help="Trabaja sobre '%s' y no toca produccion." % EXCEL_TEST)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--publicar-area", default="",
                    help="Marca Publicar=SI a todo lo que esta sin decidir en "
                         "esas Areas (coma). Ej: --publicar-area COMPRAS,JEFES. "
                         "Nunca pisa un NO que pusiste vos.")
    args = ap.parse_args()

    with open(CONFIG, encoding="utf-8") as f:
        cfg = json.load(f)
    mapa_areas = cfg["areas"]
    agr_def = cfg.get("agrupacion_default", {})
    alias = {norm(k): v for k, v in cfg.get("alias_series", {}).items()}
    ignorar = [norm(x) for x in cfg.get("ignorar", [])]
    descartar = {x.strip().lower() for x in cfg.get("descartar_reportid", [])}
    # {nombre del link a mano: agrupacion en la que se publica}
    rep_cfg = cfg.get("replicar_en_todas_las_areas", {})
    if isinstance(rep_cfg, list):          # compatibilidad con la version vieja
        rep_cfg = {x: cfg.get("agrupacion_apps_replicadas", "HERRAMIENTAS")
                   for x in rep_cfg}
    replicar = {norm(k): v for k, v in rep_cfg.items()}
    excluir_upn = tuple(norm(x) for x in cfg.get("excluir_usuarios", []) if str(x).strip())
    extra_embed = cfg.get("parametros_embed", "")
    como_informe = {norm(x) for x in cfg.get("tratar_como_informe", [])}
    reglas_agr = [(norm(k), v) for k, v in cfg.get("agrupacion_por_nombre", {}).items()]
    meses_max = int(cfg.get("meses_max", 12))  # 0 = sin tope
    abrir_areas = {x.strip().upper() for x in args.publicar_area.split(",") if x.strip()}
    # "SI" (default): el Excel replica las areas de trabajo, todo se publica y el
    # Catalogo sirve para EXCLUIR con NO. "" = al reves, hay que tildar cada uno.
    todo_por_defecto = norm(cfg.get("publicar_nuevos", "SI")) == "si"

    # --test redirige todo al Excel de prueba, sin tocar produccion
    if args.test:
        args.entrada = EXCEL_TEST if os.path.isfile(
            os.path.join(AQUI, EXCEL_TEST)) else EXCEL
        args.salida = EXCEL_TEST
    entrada = os.path.join(AQUI, args.entrada)
    if not os.path.isfile(entrada):
        entrada = os.path.join(AQUI, EXCEL)
    salida = os.path.join(AQUI, args.salida)
    print("Entrada: %s" % os.path.basename(entrada))
    print("Salida : %s%s" % (os.path.basename(salida),
                             "  (DRY-RUN, no se escribe)" if args.dry_run else ""))

    prev_tab = leer_tableros_legacy(entrada)
    prev_cat = leer_hoja(entrada, ["Catalogo"], COLS_CATALOGO)
    print("Excel actual: %d filas en Tableros, %d en Catalogo"
          % (len(prev_tab), len(prev_cat)))

    # ---- 1. curaduria previa, indexada por reportId
    curado = {}       # reportId -> dict con Area/Agrupacion/Nombre/Descripcion
    publicar = {}     # reportId -> "SI" / "NO" / ""
    for f in prev_cat:
        rid = (f["reportId"] or "").lower()
        if not rid:
            continue
        curado[rid] = f
        publicar[rid] = norm(f["Publicar"]).upper()

    manuales = []
    huerfanos = []
    for f in prev_tab:
        rid = id_de_link(f["Link"])
        if rid and norm(f["Origen"]) != "manual":
            # fila de Power BI: nombre/agrupacion/desc son curaduria a preservar
            curado.setdefault(rid, f)
            if not publicar.get(rid):
                publicar[rid] = "SI"
        else:
            f["Origen"] = "MANUAL"
            # Tipo: lo que ya pusiste manda; si no, lo deduce del link salvo que
            # este declarado como informe en tratar_como_informe.
            if norm(f.get("Tipo")) not in ("informe", "app"):
                f["Tipo"] = ("INFORME" if (rid or norm(f["Nombre"]) in como_informe)
                             else "APP")
            else:
                f["Tipo"] = norm(f["Tipo"]).upper()
            manuales.append(f)

    # ---- 1b. apps que van en todas las areas de trabajo: una sola fila de
    #      origen en el Excel, replicada. Si cambia la URL, se edita una vez.
    if replicar:
        areas_trabajo = list(dict.fromkeys(mapa_areas.values()))
        ya = {(f["Area"], norm(f["Nombre"])) for f in manuales}
        # todas las copias (y el original) van juntas al final del area, para
        # que las tarjetas de arriba sean siempre los informes
        for f in manuales:
            if norm(f["Nombre"]) in replicar:
                f["Agrupacion"] = replicar[norm(f["Nombre"])]
        copias = []
        for f in manuales:
            if norm(f["Nombre"]) not in replicar:
                continue
            for ar in areas_trabajo:
                if (ar, norm(f["Nombre"])) in ya:
                    continue
                nueva = dict(f)
                nueva["Area"] = ar
                nueva["Agrupacion"] = replicar[norm(f["Nombre"])]
                copias.append(nueva)
                ya.add((ar, norm(f["Nombre"])))
        if copias:
            manuales.extend(copias)
            print("Apps replicadas en las areas de trabajo: %d filas" % len(copias))

    # ---- 2. traer todo de la API
    print("\nConsultando la API de Power BI...")
    try:
        grupos = api("/groups")["value"]
    except NoLogueado as e:
        print("\nERROR: no hay sesion de Power BI valida (%s)." % e)
        print("Abri el orquestador y volve a loguear con el boton de Power BI.")
        return 2
    encontrados = {g["name"]: g for g in grupos}
    faltan = [w for w in mapa_areas if w not in encontrados]
    if faltan:
        print("  AVISO: no veo estas areas de trabajo (revisa permisos o el nombre): %s"
              % ", ".join(faltan))

    descubierto = []
    con_uso, sin_uso, descartadas = [], [], 0
    for nombre_ws, area in mapa_areas.items():
        g = encontrados.get(nombre_ws)
        if not g:
            continue
        reps = api("/groups/%s/reports" % g["id"])["value"]
        carpetas = carpetas_por_reporte(g["id"])
        vistas, fuera_cuenta = vistas_por_reporte(g["id"], excluir_upn)
        descartadas += fuera_cuenta
        sin_uso.append(nombre_ws) if not vistas else con_uso.append(nombre_ws)
        usados = 0
        por_carpeta = 0
        for r in reps:
            nom = (r.get("name") or "").strip()
            if not nom or any(x in norm(nom) for x in ignorar):
                continue
            rid = (r["id"] or "").lower()
            carpeta = carpetas.get(rid, "")
            serie, per = partir_serie(nom, alias, carpeta)
            if per and partir_periodo(nom)[1] is None:
                por_carpeta += 1
            descubierto.append({
                "reportId": rid,
                "Carpeta": carpeta,
                "AreaDeTrabajo": nombre_ws,
                "Area": area,
                "NombrePowerBI": nom,
                "Serie": serie or "",
                "Mes": ("%02d-%04d" % (per[1], per[0])) if per else "",
                "Vistas": vistas.get(rid, 0),
                "_orden": (per or (0, 0)),
            })
            usados += 1
        extra = ""
        if not carpetas:
            extra = "  (sin datos de subcarpetas)"
        elif por_carpeta:
            extra = "  (%d mensuales detectados por la carpeta)" % por_carpeta
        print("  %-26s -> %-11s %3d reportes%s" % (nombre_ws, area, usados, extra))

    vivos = {d["reportId"] for d in descubierto}

    # ---- 3. area efectiva: la que ya elegiste manda sobre la del workspace
    #        (ej. COMPARATIVO 08-2026 vive en Jefes pero vos lo pusiste en DIRECTORIO)
    for d in descubierto:
        cur = curado.get(d["reportId"], {})
        d["Area"] = cur.get("Area") or d["Area"]

    # ---- 4. herencia de Publicar dentro de una serie, con tope de meses
    series_si = set()
    for d in descubierto:
        if d["Serie"] and publicar.get(d["reportId"]) == "SI":
            series_si.add((d["Area"], d["Serie"]))

    # todo mes de una serie ya publicada entra como SI; el tope de meses no se
    # aplica aca sino al publicar, asi cambiar meses_max no obliga a re-tildar
    heredables = {d["reportId"] for d in descubierto
                  if d["Serie"] and (d["Area"], d["Serie"]) in series_si}

    # ---- 5. armar Catalogo
    catalogo = []
    nuevos_ids = set()
    for d in sorted(descubierto,
                    key=lambda x: (x["Area"], x["Serie"] or x["NombrePowerBI"],
                                   x["_orden"], x["NombrePowerBI"])):
        rid = d["reportId"]
        cur = curado.get(rid, {})
        pub = publicar.get(rid)
        if pub is None:
            nuevos_ids.add(rid)
        if norm(pub) == "no":
            pub = "NO"           # lo excluiste vos: se respeta siempre
        elif todo_por_defecto:
            pub = "SI"           # replicar el area de trabajo
        elif not pub:
            pub = "SI" if (rid in heredables
                           or d["Area"].upper() in abrir_areas) else ""
        if d["Serie"]:
            nom = nombre_publicado(d["Serie"], (int(d["Mes"][3:]), int(d["Mes"][:2])))
        else:
            nom = cur.get("Nombre") or d["NombrePowerBI"].upper()
        catalogo.append({
            "Publicar": pub,
            "Area": cur.get("Area") or d["Area"],
            "Agrupacion": (cur.get("Agrupacion")
                           or agrupacion_de(d["NombrePowerBI"], reglas_agr)
                           or agr_def.get(d["AreaDeTrabajo"], "OTROS")),
            "Nombre": nom,
            "Descripcion": cur.get("Descripcion", ""),
            "Serie": d["Serie"],
            "Mes": d["Mes"],
            "NombrePowerBI": d["NombrePowerBI"],
            "Carpeta": d.get("Carpeta", ""),
            "Vistas": d.get("Vistas", 0),
            "AreaDeTrabajo": d["AreaDeTrabajo"],
            "reportId": rid,
        })

    # ---- 6. links de Power BI que estaban en el Excel y ya no estan en la API
    # Si borraste el informe de Power BI, sale del tablero. No se rescata: el
    # Excel replica lo que hay en las areas de trabajo. Los links MANUAL no
    # pasan por aca, asi que los de Vercel no se ven afectados.
    for rid, cur in curado.items():
        if rid in vivos or rid in descartar:
            continue
        huerfanos.append((rid, cur.get("Nombre", "?"), cur.get("Area", "?")))

    # ---- 7. armar Tableros
    publicados = [f for f in catalogo if norm(f["Publicar"]) == "si"]

    # Un mismo informe puede estar en varias areas de trabajo (ej. COMPARATIVO
    # AGOSTO 2026 esta en Directorio y en Jefes). Si dos caen en la misma Area
    # con el mismo Nombre, se publica uno solo: gana el del area de trabajo que
    # corresponde a esa Area.
    duplicados = []
    unicos = {}
    for f in publicados:
        k = (f["Area"], f["Nombre"])
        nativo = 0 if mapa_areas.get(f["AreaDeTrabajo"]) == f["Area"] else 1
        if k not in unicos:
            unicos[k] = (nativo, f)
        elif nativo < unicos[k][0]:
            duplicados.append(unicos[k][1])
            unicos[k] = (nativo, f)
        else:
            duplicados.append(f)
    publicados = [v[1] for v in unicos.values()]

    # tope de meses: de cada serie bajan al tablero solo los mas recientes.
    # El resto queda en SI en el Catalogo (no se pierde el tilde) pero no se
    # publica, asi el desplegable de meses no se satura.
    recortados = []
    if meses_max > 0:
        por_serie = {}
        sueltos = []
        for f in publicados:
            if f["Serie"] and f["Mes"]:
                por_serie.setdefault((f["Area"], f["Serie"]), []).append(f)
            else:
                sueltos.append(f)
        publicados = list(sueltos)
        for k, fs in por_serie.items():
            fs.sort(key=lambda x: (int(x["Mes"][3:]), int(x["Mes"][:2])), reverse=True)
            publicados.extend(fs[:meses_max])
            recortados.extend(fs[meses_max:])
        publicados.sort(key=lambda x: (x["Area"], x["Agrupacion"], x["Nombre"]))

    # Las series mensuales se muestran como una sola tarjeta: para que la tarjeta
    # se ordene bien, todos sus meses llevan el total de la serie.
    total_serie = {}
    for f in publicados:
        if f["Serie"]:
            k = (f["Area"], f["Serie"])
            total_serie[k] = total_serie.get(k, 0) + int(f.get("Vistas") or 0)
    for f in publicados:
        f["_uso"] = (total_serie[(f["Area"], f["Serie"])] if f["Serie"]
                     else int(f.get("Vistas") or 0))

    # mas usado primero; lo que no tiene dato de uso queda al final, alfabetico
    publicados.sort(key=lambda f: (f["Area"], f["Agrupacion"], -f["_uso"], f["Nombre"]))

    api_rows = [{
        "Area": f["Area"], "Agrupacion": f["Agrupacion"], "Nombre": f["Nombre"],
        "Link": link_de(f["reportId"], extra_embed),
        "Descripcion": f["Descripcion"],
        "Origen": "API", "Tipo": "INFORME", "Vistas": f["_uso"],
    } for f in publicados]
    tableros = manuales + api_rows

    # ---- 8. informe
    print("\n--- Resumen ---")
    print("  Reportes vistos en Power BI ....... %d" % len(descubierto))
    print("  Publicados (Publicar=SI) ......... %d" % len(publicados))
    print("  Links a mano preservados ......... %d" % len(manuales))
    print("  Areas con datos de uso ........... %s" % (", ".join(con_uso) or "ninguna"))
    if excluir_upn:
        print("  Vistas descartadas (%s) ... %d"
              % (", ".join(excluir_upn), descartadas))
    if sin_uso:
        print("  SIN datos de uso (orden alfabetico): %s" % ", ".join(sin_uso))
        print("    Para activarlos: en Power BI Service, abri un informe de esa area")
        print("    de trabajo > Metricas de uso. Power BI crea el modelo y la")
        print("    proxima corrida ya ordena por uso.")
    print("  Nuevos desde la ultima corrida ... %d" % len(nuevos_ids))
    if duplicados:
        print("  Repetidos en 2 areas de trabajo, publicado 1 ... %d" % len(duplicados))
        for f in duplicados:
            print("      %-11s %-24s se descarto el de %s"
                  % (f["Area"], f["Nombre"], f["AreaDeTrabajo"]))
    if recortados:
        series = sorted({(f["Area"], f["Serie"]) for f in recortados})
        print("  Meses viejos NO publicados (tope meses_max=%d) ... %d"
              % (meses_max, len(recortados)))
        for ar, se in series:
            ms = sorted((f["Mes"] for f in recortados
                         if (f["Area"], f["Serie"]) == (ar, se)))
            print("      %-11s %-24s %s" % (ar, se, ", ".join(ms)))
    auto = [f for f in catalogo
            if f["reportId"] in nuevos_ids and norm(f["Publicar"]) == "si"]
    if auto:
        print("\n  Meses nuevos publicados solos (heredaron la serie):")
        for f in auto:
            print("    + %-11s %-32s  <- %s" % (f["Area"], f["Nombre"], f["NombrePowerBI"]))
    pend = [f for f in catalogo
            if f["reportId"] in nuevos_ids and norm(f["Publicar"]) != "si"]
    if pend:
        print("\n  Nuevos SIN publicar (marca Publicar=SI en la hoja Catalogo si los queres):")
        for f in pend[:40]:
            print("    . %-11s %s" % (f["Area"], f["NombrePowerBI"]))
        if len(pend) > 40:
            print("    ... y %d mas" % (len(pend) - 40))
    if huerfanos:
        print("")
        print("  Borrados de Power BI, salieron del tablero:")
        for rid, nom, ar in huerfanos:
            print("    ! %-11s %-34s %s" % (ar, nom, rid))

    if args.dry_run:
        print("\nDRY-RUN: no escribi nada.")
        return 0
    try:
        escribir(salida, tableros, catalogo, nuevos_ids)
    except PermissionError:
        print("")
        print("ERROR: no puedo escribir '%s'." % os.path.basename(salida))
        print("  Casi seguro lo tenes abierto en Excel. Cerralo y volve a correr.")
        print("  (No se modifico nada: el tablero quedo como estaba.)")
        return 3
    print("\nListo: %s  (%d filas en Tableros, %d en Catalogo)"
          % (os.path.basename(salida), len(tableros), len(catalogo)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
